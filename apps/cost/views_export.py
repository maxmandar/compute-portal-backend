from rest_framework.views import APIView
from rest_framework.parsers import JSONParser
from rest_framework.response import Response
from django.http import HttpResponse
import openpyxl
from openpyxl import Workbook


from django.http import HttpResponse

def generate_excel(payload):
    wb = Workbook()
    ws = wb.active

    # Add header row
    header_row = ["Item", "Configuration", "Category", "Cost", "Type"]
    for col_num, header in enumerate(header_row, 1):
        ws.cell(row=1, column=col_num, value=header)

    # Add data rows
    for row_num, data in enumerate(payload, 2):
        ws.cell(row=row_num, column=1, value=data["item"])
        ws.cell(row=row_num, column=2, value=data["config"])
        ws.cell(row=row_num, column=3, value=data["catagory"])
        ws.cell(row=row_num, column=4, value=data["cost"])
        ws.cell(row=row_num, column=5, value=data["type"])

    return wb

class ExportExcelView(APIView):
    parser_classes = [JSONParser]

    def post(self, request, *args, **kwargs):
        payload = request.data
        wb = generate_excel(payload)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=vmware_server_cost.xlsx'
        wb.save(response)

        return response